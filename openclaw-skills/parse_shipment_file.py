"""
OpenClaw Skill: parse_shipment_file
解析客户发来的 Excel/PDF 文件，提取结构化运输数据，然后调用 EW-WebView API 创建订单。

Supported formats:
  - .xlsx / .xls  (Excel)
  - .pdf           (PDF with tabular data)
  - .csv           (CSV)

Usage in OpenClaw:
  This skill is triggered via webhook when WeCom receives a file from a customer group.
  OpenClaw downloads the file, runs this skill, and posts results to the EW-WebView agent API.
"""

import json
import os
import re
import sys
from pathlib import Path

# --- Configuration ---
EW_API_BASE = os.environ.get("EW_API_BASE_URL", "http://localhost:5001/api")
AGENT_API_KEY = os.environ.get("AGENT_WEBHOOK_API_KEY", "")


def extract_excel(file_path: str) -> list[list]:
    """Extract all rows from an Excel file as lists of cell values."""
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl")
        import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
    return rows


def extract_pdf(file_path: str) -> list[list]:
    """Extract tabular data from a PDF file."""
    try:
        import pdfplumber
    except ImportError:
        os.system("pip install pdfplumber")
        import pdfplumber

    rows = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                rows.extend(table)
            if not tables:
                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        rows.append(line.split())
    return rows


def extract_csv(file_path: str) -> list[list]:
    """Extract rows from a CSV file."""
    import csv
    rows = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    return rows


# --- Column Mapping ---
# Maps common Chinese/English header variants to our standard field names
HEADER_MAP = {
    "包装类型": "packaging_type",
    "外箱单号": "tracking_number",
    "备注": "notes",
    "派送方式": "delivery_method",
    "中文品名": "product_name_cn",
    "英文品名": "product_name_en",
    "价值": "cargo_value",
    "国家中文": "_country_cn",
    "国家英文": "destination_country",
    "邮编": "destination_zip",
    "城市": "destination_city",
    "公司名": "company_name",
    "收件人姓名": "recipient_name",
    "收件人": "recipient_name",
    "电话": "phone",
    "邮箱": "email",
    "详细地址": "address",
    "地址": "address",
    "箱数": "total_pieces",
    "尺寸箱数": "_dim_pieces",
    "实重": "_weight",
    "方数": "_volume",
    "长（CM）": "_length",
    "长(CM)": "_length",
    "长": "_length",
    "宽（CM）": "_width",
    "宽(CM)": "_width",
    "宽": "_width",
    "高（CM）": "_height",
    "高(CM)": "_height",
    "高": "_height",
    "报价（美金）": "_quote_ref",
    "入仓备注": "_warehouse_notes",
    "地址类型": "address_type",
}


def map_headers(header_row: list) -> dict:
    """Map the header row to standard field names. Returns {col_index: field_name}."""
    mapping = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip()
        if cell_str in HEADER_MAP:
            mapping[i] = HEADER_MAP[cell_str]
        else:
            for key, field in HEADER_MAP.items():
                if key in cell_str or cell_str in key:
                    mapping[i] = field
                    break
    return mapping


def parse_rows_to_items(rows: list[list]) -> list[dict]:
    """
    Parse extracted rows into structured shipment items.
    Handles the "main row + sub-rows for dimensions" pattern.
    """
    if not rows:
        return []

    header_mapping = map_headers(rows[0])
    if not header_mapping:
        return []

    items = []
    current_item = None

    for row_idx, row in enumerate(rows[1:], start=1):
        row_dict = {}
        for col_idx, field in header_mapping.items():
            if col_idx < len(row):
                row_dict[field] = row[col_idx]

        tracking = row_dict.get("tracking_number")
        has_main_info = tracking or row_dict.get("product_name_cn") or row_dict.get("product_name_en")

        dim = None
        weight = row_dict.get("_weight")
        length = row_dict.get("_length")
        width = row_dict.get("_width")
        height = row_dict.get("_height")
        volume = row_dict.get("_volume")
        dim_pieces = row_dict.get("_dim_pieces")

        if any(v is not None for v in [weight, length, width, height]):
            dim = {
                "pieces": _to_int(dim_pieces, 1),
                "weight": _to_float(weight),
                "length": _to_float(length),
                "width": _to_float(width),
                "height": _to_float(height),
                "volume": _to_float(volume),
            }

        if has_main_info:
            if current_item:
                items.append(current_item)

            # Combine address_type from both 地址类型 and 入仓备注 columns
            addr_type = _clean_str(row_dict.get("address_type")) or ""
            warehouse_notes = _clean_str(row_dict.get("_warehouse_notes")) or ""
            combined_addr = f"{warehouse_notes} {addr_type}".strip() if (warehouse_notes or addr_type) else ""

            current_item = {
                "tracking_number": _clean_str(tracking),
                "packaging_type": _clean_str(row_dict.get("packaging_type")),
                "product_name_cn": _clean_str(row_dict.get("product_name_cn")),
                "product_name_en": _clean_str(row_dict.get("product_name_en")),
                "cargo_value": _to_float(row_dict.get("cargo_value")),
                "destination_country": _clean_str(row_dict.get("destination_country")) or "US",
                "destination_zip": _clean_str(row_dict.get("destination_zip")),
                "destination_city": _clean_str(row_dict.get("destination_city")),
                "company_name": _clean_str(row_dict.get("company_name")),
                "recipient_name": _clean_str(row_dict.get("recipient_name")),
                "phone": _clean_str(row_dict.get("phone")),
                "email": _clean_str(row_dict.get("email")),
                "address": _clean_str(row_dict.get("address")),
                "address_type": combined_addr,
                "total_pieces": _to_int(row_dict.get("total_pieces"), 1),
                "delivery_method": _clean_str(row_dict.get("delivery_method")),
                "notes": _clean_str(row_dict.get("notes")),
                "dimensions": [],
            }
            if dim:
                current_item["dimensions"].append(dim)
        elif current_item and dim:
            current_item["dimensions"].append(dim)

    if current_item:
        items.append(current_item)

    return items


def _clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _to_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        cleaned = re.sub(r"[^\d.]", "", str(val))
        try:
            return float(cleaned) if cleaned else default
        except ValueError:
            return default


def _to_int(val, default=1):
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def post_to_ew_api(items: list[dict], source_file: str, wecom_chat_id: str = None):
    """Post parsed items to EW-WebView agent webhook API."""
    try:
        import requests
    except ImportError:
        os.system("pip install requests")
        import requests

    payload = {
        "items": items,
        "sourceFile": source_file,
        "wecomChatId": wecom_chat_id,
    }

    headers = {"Content-Type": "application/json"}
    if AGENT_API_KEY:
        headers["X-Agent-Api-Key"] = AGENT_API_KEY

    resp = requests.post(
        f"{EW_API_BASE}/agent/webhook",
        json=payload,
        headers=headers,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def parse_file(file_path: str) -> list[dict]:
    """Main entry point: parse a file and return structured items."""
    ext = Path(file_path).suffix.lower()

    if ext in (".xlsx", ".xls"):
        rows = extract_excel(file_path)
    elif ext == ".pdf":
        rows = extract_pdf(file_path)
    elif ext == ".csv":
        rows = extract_csv(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    items = parse_rows_to_items(rows)
    return items


# --- CLI entry point for testing ---
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_shipment_file.py <file_path> [--post]")
        sys.exit(1)

    file_path = sys.argv[1]
    should_post = "--post" in sys.argv

    items = parse_file(file_path)
    print(f"\nParsed {len(items)} shipment items:\n")
    print(json.dumps(items, indent=2, ensure_ascii=False))

    if should_post and items:
        result = post_to_ew_api(items, os.path.basename(file_path))
        print(f"\nPosted to EW API: {json.dumps(result, indent=2)}")
